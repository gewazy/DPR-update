import pyodbc
import openpyxl
from datetime import date, timedelta
from time import sleep

# do zrobienia:
# - przepisać skrypt w kilka modułów by był łatwiejszy w edycji
#   /kwerendy w oddzielnym pliku, scieżki w oddzielnym pliku/
# - rozróżnianie zmian od niwelacji juz po podegraniu remeasuringu do postplotu.
#   /sprawdz czy nazwisko jest w remeasuringu, od liczb w zmianach odjac liczby z remeasuring/
# - rozpoznie ostatniego dnia w raporcie i przygotowanie xlsx do dnia obecnego, oraz wypełnienie dni po kolei
#   /jesli oststnia zakladka jest starsza niż z dnia dzisiejszego iterowac przez skrypt dla poszczegolnych dni


print('Generowania raportu DPR.\n\tver.beta')

print(f"-dane z dnia: {str(date.today().strftime('%Y%m%d'))}")

# łączenie z bazą
conn_str = (r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};' # sterownik mdb
            r'DBQ=..\01_database\PL-182 HUSOW.mdb;')  # ścieżka do bazy danych

cnxn = pyodbc.connect(conn_str)  # łączenie z bazą danych
crsr = cnxn.cursor()

# kwerendy - jak nazrazie trzeba zmieniać z palucha
tycz_r = "Select " \
         "[Ludziki].`Nr_auta`,  " \
         "'1' as `L_B`, " \
         "IIF (`Survey Mode (value)`=5, 'ZUPT ', IIF (`Survey Mode (value)`=6,'TACHIMETR', IIF (`Survey Mode (value)` in (3, 10, 1, 2, 13, 9),'GPS',''))) & ' ' & [POSTPLOT].`Surveyor` AS `Brygada`, " \
         "Count (*) AS `Liczba PW` " \
         "From " \
         "[POSTPLOT] Left Join [Ludziki] on [POSTPLOT].`Surveyor`=[Ludziki].`Surveyor` " \
         "Where " \
         "[POSTPLOT].`Offset (North)` is not NULL " \
         "and `IsDuplicate` is NULL " \
         "And [POSTPLOT].`Station (value)` > 0 " \
         "And [POSTPLOT].`Track` Between 1175 And 1930 " \
         "And datediff('d',[POSTPLOT].`Survey Time (Local)`,Now()) = 0 " \
         "Group By " \
         "IIF (`Survey Mode (value)`=5, 'ZUPT ', IIF (`Survey Mode (value)`=6, 'TACHIMETR', IIF (`Survey Mode (value)` in (3, 10, 1, 2, 13, 9),'GPS',''))), " \
         "[POSTPLOT].`Surveyor`, " \
         "[POSTPLOT].`Julian Date (Local)`, " \
         "[Ludziki].`Nr_auta`"

tycz_s = "Select " \
         "[Ludziki].`Nr_auta`, " \
         " '1' as `L_B`, " \
         "IIF (`Survey Mode (value)`=5, 'ZUPT ', IIF (`Survey Mode (value)`=6,'TACHIMETR', IIF (`Survey Mode (value)` in (3, 10, 1, 2, 13, 9),'GPS',''))) & ' ' & [POSTPLOT].`Surveyor` AS `Brygada`, " \
         "Count (*) AS `Liczba PW` " \
         "From " \
         "[POSTPLOT] Left Join [Ludziki] on [POSTPLOT].`Surveyor`=[Ludziki].`Surveyor` " \
         "Where " \
         "[POSTPLOT].`Offset (North)` is not NULL " \
         "and `IsDuplicate` is NULL " \
         "And [POSTPLOT].`Station (value)` > 0 " \
         "And  [POSTPLOT].`Track` Between 4060 And 4550 " \
         "And datediff ('d',[POSTPLOT].`Survey Time (Local)`,Now()) = 0 " \
         "Group By " \
         "IIF (`Survey Mode (value)`=5, 'ZUPT ', IIF (`Survey Mode (value)`=6, 'TACHIMETR', IIF (`Survey Mode (value)` in (3, 10, 1, 2, 13, 9),'GPS',''))), " \
         "[POSTPLOT].`Surveyor`, " \
         "[POSTPLOT].`Julian Date (Local)`, " \
         "[Ludziki].`Nr_auta`"

zm_r = "Select " \
       "[Ludziki].`Nr_auta`,  " \
       "'1' as `L_B`, " \
       "IIF (`Survey Mode (value)`=5, 'ZUPT ', IIF (`Survey Mode (value)`=6,'TACHIMETR', IIF (`Survey Mode (value)` in (3, 10, 1, 2, 13, 9),'GPS',''))) & ' ' & [POSTPLOT].`Surveyor` AS `Brygada`, " \
       "Count (*) AS `Liczba PW` " \
       "From " \
       "[POSTPLOT] Left Join [Ludziki] on [POSTPLOT].`Surveyor`=[Ludziki].`Surveyor` " \
       "Where " \
       "[POSTPLOT].`Offset (North)` is not NULL " \
       "and `IsDuplicate` is not NULL " \
       "And [POSTPLOT].`Station (value)` > 0 " \
       "And  [POSTPLOT].`Track` Between 1175 And 1930 " \
       "And datediff ('d',[POSTPLOT].`Survey Time (Local)`,Now()) = 0 " \
       "Group By " \
       "IIF (`Survey Mode (value)`=5, 'ZUPT ', IIF (`Survey Mode (value)`=6, 'TACHIMETR', IIF (`Survey Mode (value)` in (3, 10, 1, 2, 13, 9),'GPS',''))), " \
       "[POSTPLOT].`Surveyor`, " \
       "[POSTPLOT].`Julian Date (Local)`, " \
       "[Ludziki].`Nr_auta`"

zm_s = "Select " \
       "[Ludziki].`Nr_auta`,  " \
       "'1' as `L_B`, " \
       "IIF (`Survey Mode (value)`=5, 'ZUPT ', IIF (`Survey Mode (value)`=6,'TACHIMETR', IIF (`Survey Mode (value)` in (3, 10, 1, 2, 13, 9),'GPS',''))) & ' ' & [POSTPLOT].`Surveyor` AS `Brygada`, " \
       "Count (*) AS `Liczba PW` " \
       "From " \
       "[POSTPLOT] Left Join [Ludziki] on [POSTPLOT].`Surveyor`=[Ludziki].`Surveyor` " \
       "Where " \
       "[POSTPLOT].`Offset (North)` is not NULL " \
       "and `IsDuplicate` is not NULL " \
       "And [POSTPLOT].`Station (value)` > 0 " \
       "And  [POSTPLOT].`Track` Between 4060 And 4550 " \
       "And datediff ('d',[POSTPLOT].`Survey Time (Local)`,Now()) = 0 " \
       "Group By " \
       "IIF (`Survey Mode (value)`=5, 'ZUPT ', IIF (`Survey Mode (value)`=6, 'TACHIMETR', IIF (`Survey Mode (value)` in (3, 10, 1, 2, 13, 9),'GPS',''))), " \
       "[POSTPLOT].`Surveyor`, " \
       "[POSTPLOT].`Julian Date (Local)`, " \
       "[Ludziki].`Nr_auta`"

re_s = "Select " \
       "[Ludziki].`Nr_auta`, " \
       " '1' as `L_B`, " \
       "IIF (`Survey Mode (value)`=5, 'ZUPT ', IIF (`Survey Mode (value)`=6,'TACHIMETR', IIF (`Survey Mode (value)` in (3, 10, 1, 2, 13, 9),'GPS',''))) & ' ' & [REMEASURE].`Surveyor` AS `Brygada`, " \
       "Count (*) AS `Liczba PW` " \
       "From " \
       "[REMEASURE] Left Join [Ludziki] on [REMEASURE].`Surveyor`=[Ludziki].`Surveyor` " \
       "Where " \
       "[REMEASURE].`Offset (North)` is not NULL " \
       "and `IsDuplicate` is NULL " \
       "And [REMEASURE].`Station (value)` > 0 " \
       "And  [REMEASURE].`Track` Between 4060 And 4550 " \
       "And datediff ('d',[REMEASURE].`Survey Time (Local)`,Now()) = 0 " \
       "Group By " \
       "IIF (`Survey Mode (value)`=5, 'ZUPT ', IIF (`Survey Mode (value)`=6, 'TACHIMETR', IIF (`Survey Mode (value)` in (3, 10, 1, 2, 13, 9),'GPS',''))), " \
       "[REMEASURE].`Surveyor`, " \
       "[REMEASURE].`Julian Date (Local)`, " \
       "[Ludziki].`Nr_auta`"

re_r = "Select " \
       "[Ludziki].`Nr_auta`,  " \
       "'1' as `L_B`, " \
       "IIF (`Survey Mode (value)`=5, 'ZUPT ', IIF (`Survey Mode (value)`=6,'TACHIMETR', IIF (`Survey Mode (value)` in (3, 10, 1, 2, 13, 9),'GPS',''))) & ' ' & [REMEASURE].`Surveyor` AS `Brygada`, " \
       "Count (*) AS `Liczba PW` " \
       "From " \
       "[REMEASURE] Left Join [Ludziki] on [REMEASURE].`Surveyor`=[Ludziki].`Surveyor` " \
       "Where " \
       "[REMEASURE].`Offset (North)` is not NULL " \
       "and `IsDuplicate` is NULL " \
       "And [REMEASURE].`Station (value)` > 0 " \
       "And [REMEASURE].`Track` Between 1175 And 1930 " \
       "And datediff('d',[REMEASURE].`Survey Time (Local)`,Now()) = 0 " \
       "Group By " \
       "IIF (`Survey Mode (value)`=5, 'ZUPT ', IIF (`Survey Mode (value)`=6, 'TACHIMETR', IIF (`Survey Mode (value)` in (3, 10, 1, 2, 13, 9),'GPS',''))), " \
       "[REMEASURE].`Surveyor`, " \
       "[REMEASURE].`Julian Date (Local)`, " \
       "[Ludziki].`Nr_auta`"


qc_r = "Select [POSTPLOT].* " \
       "From [POSTPLOT] " \
       "Where  [POSTPLOT].`Station (value)` > 0 And [POSTPLOT].`Track` Between 1175 And 1930  " \
       "And [POSTPLOT].`Status` >=1 And [POSTPLOT].`Status` <= 11 " \
       "And (( [POSTPLOT].`Survey Mode (value)` Not In (3,5,6) ) Or ( [POSTPLOT].`Survey Mode (value)` = 3 And ([POSTPLOT].`Number of Satellites` < 5 Or [POSTPLOT].`PDOP` > 6 Or [POSTPLOT].`CQ` > 0.3) ))  " \
       "Order By [POSTPLOT].`Station (text)`"

qc_s = "Select [POSTPLOT].* " \
       "From [POSTPLOT] " \
       "Where [POSTPLOT].`Station (value)` > 0 And [POSTPLOT].`Track` Between 4060 And 4550  " \
       "And ( ([POSTPLOT].`Status` IN (2,4) And  [POSTPLOT].`Survey Mode (value)` Not In (3,5,6))  Or ( [POSTPLOT].`Status` = 5  And  [POSTPLOT].`Survey Mode (value)` In (3) And ([POSTPLOT].`Number of Satellites` < 5 Or [POSTPLOT].`PDOP` > 6 Or [POSTPLOT].`CQ` > 0.3) )  or [POSTPLOT].`Status` = 5 or [POSTPLOT].`Status` = 6 ) " \
       "Order By [POSTPLOT].`Station (value)`"

vib = "Select [POSTPLOT].* From [POSTPLOT] " \
      "Where  [POSTPLOT].`Status` <> 0 And  [POSTPLOT].`Track` Between 4060 And 4550  And [POSTPLOT].`Station (value)`>0 " \
      "And (([POSTPLOT].`Descriptor` in ('x40', 'x45', 'x30', 'x35', 'x20', 'x25', 'x10', 'x15', 'xm', 'xm5') OR (([POSTPLOT].`Descriptor` Like 'xt' Or [POSTPLOT].`Descriptor` Like 'xr') and [POSTPLOT].`Status`  in (3,4,5)))) " \
      "Order By [POSTPLOT].`Station (value)`"

xr = "Select [POSTPLOT].* From [POSTPLOT] Where [POSTPLOT].`Status` <> 0 And [POSTPLOT].`Track` Between 4060 And 4550  " \
     "And [POSTPLOT].`Station (value)`<>0 And (([POSTPLOT].`Descriptor` Like 'xr' And [POSTPLOT].`dr_date` is NULL) OR ([POSTPLOT].`dr_date` is not NULL And ([POSTPLOT].`dr_eq` Like 'EMCI' Or [POSTPLOT].`dr_eq` Like  'Emci' Or [POSTPLOT].`dr_eq` Like  'LPHB'))) And [POSTPLOT].`Status` not in (3,4,5) Order By [POSTPLOT].`Station (value)`"

xt = "Select [POSTPLOT].* From [POSTPLOT] " \
     "Where [POSTPLOT].`Status` <> 0 And [POSTPLOT].`Track` Between 4060 And 4550 And [POSTPLOT].`Station (value)`<>0 And( ([POSTPLOT].`Descriptor` Like 'xt' " \
     "And [POSTPLOT].`dr_date` is NULL) OR ([POSTPLOT].`dr_date` is not NULL And ([POSTPLOT].`dr_eq` Like 'PAT' Or [POSTPLOT].`dr_eq` Like  'Pat' ))) And [POSTPLOT].`Status` not in (3,4,5) " \
     "Order By [POSTPLOT].`Station (value)`"

skip = "Select [POSTPLOT].* From [POSTPLOT] " \
       "Where [POSTPLOT].`Status` = 0 And [POSTPLOT].`Station (value)` > 0 " \
       "And [POSTPLOT].`Track` Between 4060 And 4550 Order By [POSTPLOT].`Station (value)`"


crsr.execute(vib)
vib = len(crsr.fetchall())

crsr.execute(xr)
xr = len(crsr.fetchall())

crsr.execute(xt)
xt = len(crsr.fetchall())

crsr.execute(skip)
skip = len(crsr.fetchall())

crsr.execute(qc_r)
qc_r = len(crsr.fetchall())

crsr.execute(qc_s)
qc_s = len(crsr.fetchall())

crsr.execute(tycz_r)
tycz_r = crsr.fetchall()

crsr.execute(tycz_s)
tycz_s = crsr.fetchall()

crsr.execute(zm_r)
zm_r = crsr.fetchall()

crsr.execute(zm_s)
zm_s = crsr.fetchall()

crsr.execute(re_s)
re_s = crsr.fetchall()

crsr.execute(re_r)
re_r = crsr.fetchall()

crsr.close()
cnxn.close()


# Przygotowanie pliku Excell
print('Otwieram DPR\n')
raport = '.\output\dniowki\PL_182_Raport_geodezyjny_DPR.xlsx'

wb = openpyxl.load_workbook(raport)

print('-Tworzę nową zakładkę')
source = wb[wb.sheetnames[-2]]
target = wb.copy_worksheet(source)
wb.move_sheet(target, offset=-1)

print('-Zmieniam nazwe zakładki')
target.title = str((date.today() + timedelta(days=1)).strftime('%Y%m%d'))
target.sheet_view.zoomScale = 75
target.sheet_view.view = "pageBreakPreview"

# zamiana formuł w pliku
print('-Edytuję formuły')
a = f"'{str((date.today() - timedelta(days=1)).strftime('%Y%m%d'))}'"
b = f"'{source.title}'"
for row in target.iter_rows(min_row=1, max_col=15, max_row=133):
    for cell in row:
        if type(cell.value) is str:
            cell.value = cell.value.replace(a, b)

# ustawienie arkusza aktywnego
print('\nPrzechodzę na arkusz z dnia dzisiejszego')
target = wb[wb.sheetnames[-3]]

# wypełnianie raportu
print("-Wypełniam arkusz\n")

print('--Wibratory: ', vib)
target['E132'] = vib

print('--Wiercenie ręczne: ', xr)
target['F132'] = xr

print('--Wiercenie traktorem: ', xt)
target['G132'] = xt

print('--Skipy: ', skip)
target['K132'] = skip

print('--QC R: ', qc_r)
target['G53'] = qc_r

print('--QC S: ', qc_s)
target['O53'] = qc_s

licz_bry = []  # lista brygadzistó do liczenia liczby brygad

sleep(0.7)
row = 13
print('\nTyczenie punktów odbioru: \n')
for ro in tycz_r:
    row += 1
    print(ro)
    target['A' + str(row)] = ro[0]
    target['B' + str(row)] = ro[1]
    target['C' + str(row)] = ro[2]
    licz_bry.append((ro[2].split())[1])
    target['D' + str(row)] = ro[3]
sleep(0.1)

row = 13
print('\nTyczenie punktów wzbudzania: \n')
for ro in tycz_s:
    row += 1
    print(ro)
    target['I' + str(row)] = ro[0]
    target['J' + str(row)] = ro[1]
    target['K' + str(row)] = ro[2]
    licz_bry.append((ro[2].split())[1])
    target['L' + str(row)] = ro[3]
sleep(0.1)

row = 41
print('\nDomierzanie/niwelacja punktów odbioru: \n')
for ro in re_r:
    row += 1
    print(ro)
    target['A' + str(row)] = ro[0]
    target['B' + str(row)] = ro[1]
    target['C' + str(row)] = ro[2]
    licz_bry.append((ro[2].split())[1])
    target['D' + str(row)] = ro[3]
sleep(0.1)

row = 41
print('\nDomierzanie/niwelacja punktów wzbudzania: \n')
for ro in re_s:
    row += 1
    print(ro)
    target['I' + str(row)] = ro[0]
    target['J' + str(row)] = ro[1]
    target['K' + str(row)] = ro[2]
    licz_bry.append((ro[2].split())[1])
    target['L' + str(row)] = ro[3]
sleep(0.1)

row = 57
print('\nZmiany punktów odbioru: \n')
for ro in zm_r:
    row += 1
    print(ro)
    target['A' + str(row)] = ro[0]
    target['B' + str(row)] = ro[1]
    target['C' + str(row)] = ro[2]
    licz_bry.append((ro[2].split())[1])
    target['D' + str(row)] = ro[3]
sleep(0.1)

row = 57
print('\nZmiany punktów wzbudzania: \n')
for ro in zm_s:
    row += 1
    print(ro)
    target['I' + str(row)] = ro[0]
    target['J' + str(row)] = ro[1]
    target['K' + str(row)] = ro[2]
    licz_bry.append((ro[2].split())[1])
    target['L' + str(row)] = ro[3]
sleep(0.4)

print(f'\n\nPracowało {len(set(licz_bry))} brygad\n')
for num, geodeta in enumerate(sorted(set(licz_bry))):
    print(str(num + 1) + '.' , geodeta)

target['N132'] = len(set(licz_bry))
target['B128'] = input('\nWprowadź komentarz:\n')

print('\nZapisuję plik')
wb.save(raport)

print('\nRaport DPR gotowy\n'
      '\nPrzystępuję do zrobienia raportu dziennego "dniówki"')

# raprt dniówkowy puunkt po punkcie importowany ze skryptu
import raport_daily
